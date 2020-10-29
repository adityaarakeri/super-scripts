import googletrans
from openpyxl import load_workbook
import sys


def main():
    if len(sys.argv) != 5:
        print("incorrect input parameters")
        return 1

    # Load worksheet
    worksheet = sys.argv[1]

    wb = load_workbook(worksheet)

    # Selects active worksheet (excel List) (index 0 by default == 1st List)
    ws = wb.active


    column = sys.argv[2]

    start_str = column + sys.argv[3]
    end_str = column + sys.argv[4]

    # Selects 6 cells
    rows = ws[start_str:end_str]

    # Initialize Translator Instance from Translator Object (I guess?)
    translator = googletrans.Translator()

    dest = input("Translate to (ie. cs, en, ..): ")
    src = input("Translate from: ")
    # Iterating through objects in rows
    for row in rows:
        # Uses googletrans package to ping Google Translate API (AJAX call) to translate input text
        # Input text is 1st parameter in .translate method on Translator object

        # Iterating through cells in row (It's an object)
        for cell in row:

            # Checks if cell is empty (could be checked more elegantly by removing empty cells I think)
            # If empty (doesn't have a value) skip it
            if not cell.value:
                continue
            else:
                text = translator.translate(cell.value, dest=dest, src=src).text
                cell.value = text

    # Splits on "." to get rid of the xlsx ending
    cut_postfix = worksheet.split(".")

    save_str = cut_postfix[0] + "_translated.xlsx"
    # Saves the changes to a new file

    wb.save(save_str)
    print("Successfully translated!")


if __name__ == "__main__":
    main()
