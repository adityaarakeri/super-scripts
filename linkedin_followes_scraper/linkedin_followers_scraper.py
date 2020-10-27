from selenium import webdriver
import bs4
import time
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill
from datetime import datetime

greenFont = Font(color="00FF00")
redFont = Font(color="FF0000")

browser = webdriver.Firefox()
browser.get("https://www.linkedin.com/feed/followers/")

# The page opens in SIGNUP page
signInElement = browser.find_element_by_class_name("main__sign-in-link")
signInElement.click()

# Control now in SignIn page

# password field
passEle = browser.find_element_by_id("password")
passEle.click()

# enter your password below
passEle.send_keys("your_password_here")

# username field
phoneEle = browser.find_element_by_id("username")
phoneEle.click()

# enter your username/ phone number/ email id below
phoneEle.send_keys("your_phonenumber_or_email_id_here")

# login button
loginBtn = browser.find_element_by_xpath(
    "/html/body/div/main/div[2]/form/div[4]/button"
)
loginBtn.click()

# this block of code scorolls down to thenend of the page
# make sure the page is not indefinite like a social media
# page should end somewhere
lenOfPage = browser.execute_script(
    "window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;"
)
match = False
while match == False:
    lastCount = lenOfPage
    time.sleep(2)
    lenOfPage = browser.execute_script(
        "window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;"
    )
    if lastCount == lenOfPage:
        match = True


# get the source code of page and parse it
soup = bs4.BeautifulSoup(browser.page_source, "html.parser")

# follower names are in h3 tag
followers = soup.findAll("h3")
type(followers)

# empty list to which the text i.e. name of all followers are appended
followersList = []
for follower in followers:
    followersList.append(follower.text.strip())

print("No. of followers currently= %i" % (len(followersList)))


# not really required
for person in followersList:
    print(person)


# we have follower all stored in a list now we have to add it to excels sheet
# if excel sheet exists we must get the last entries column
# to match with current entries and change font color to red or green respectively
# else create new spreadsheet

if Path("C:\\Users\\91970\\Desktop\\LinkedIn_Followers.xlsx").exists():
    # the excel sheet exists and must be having previous data
    print("Loading workbook")

    # make sure to chenge this to your desktop directory
    workbook = openpyxl.load_workbook(
        "C:\\Users\\91970\\Desktop\\LinkedIn_Followers.xlsx"
    )
    currentSheet = workbook.active

    previousList = []
    # this will give us the last column from which we have to get names
    numberOfColumns = currentSheet.max_column
    # this will give the number of entries in the last column including the date
    lengthOfLastColumn = len(currentSheet[get_column_letter(numberOfColumns)])

    print(
        "Number of columns= %d and entries in last column= %d"
        % (numberOfColumns, lengthOfLastColumn)
    )

    # starting from second row cuz first one will give the date
    for i in range(2, lengthOfLastColumn):
        previousList.append(currentSheet.cell(i, numberOfColumns).value.strip())
        if currentSheet.cell(i, numberOfColumns).value not in followersList:
            currentSheet.cell(i, numberOfColumns).font = redFont

    # print(previousList)
    # this was to check any changes in the current follower list

    rowCount = 1
    currentColumn = numberOfColumns + 1
    dateCell = currentSheet.cell(rowCount, currentColumn)
    dateCell.value = str(datetime.today().strftime("%d-%m-%Y"))

    rowCount += 1
    # add followers from secod row onwards

    for person in followersList:
        currentCell = currentSheet.cell(rowCount, currentColumn)
        currentCell.value = person.strip()
        rowCount += 1
        if person not in previousList:
            currentCell.font = greenFont

    # make sure to chenge this to your desktop directory
    workbook.save("C:\\Users\\91970\\Desktop\\LinkedIn_Followers.xlsx")
    # don't forget to save


else:
    # the excel sheet does not exist and must be created
    print("creating Workbook")
    workbook = openpyxl.Workbook()
    currentSheet = workbook.active
    # Complete everything

    rowCount = 1
    currentColumn = 1

    # add today's date in the first column
    dateCell = currentSheet.cell(rowCount, currentColumn)
    dateCell.value = str(datetime.today().strftime("%d-%m-%Y"))
    rowCount += 1

    # add people to the rest of the rows in this column
    print("adding people to sheet")
    for person in followersList:
        currentCell = currentSheet.cell(rowCount, currentColumn)
        currentCell.value = person.strip()
        rowCount += 1

    # make sure to chenge this to your desktop directory
    workbook.save("C:\\Users\\91970\\Desktop\\LinkedIn_Followers.xlsx")
    # lastly don't forget to save
